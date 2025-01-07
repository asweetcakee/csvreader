from openpyxl import Workbook, load_workbook
from datetime import datetime

class ExcelWriter:
    def __init__(self, output_file):
        self.output_file = output_file

    def write_to_excel(self, data):
        """
        :param data: Dict of processed data, key — list name, value — list of phone numbers
        """
        wb = Workbook()
        rfm_segment = self.get_rfm_date()
        sheet_created = False  # Track if at least one sheet is created

        for sheet_name, phone_numbers in data.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(["phone", "full_name", "age", "city", "import_id", "gender", "rfm_segment", "game_id"])

            # Fill in the data
            import_id = 701120240001
            for phone in phone_numbers:
                ws.append([phone, "", 1995, "", import_id, "", rfm_segment, import_id])
                import_id += 1

            sheet_created = True  # Mark that at least one sheet is created

        # Deletes the default excel sheet if it's not needed
        default_sheet = wb["Sheet"]
        if default_sheet and sheet_created:
            wb.remove(default_sheet)

        # Ensure there's at least one sheet in the workbook
        if not sheet_created:
            ws = wb.active  # Use the default sheet
            ws.title = "Empty"
            ws.append(["No data available"])  # Add a placeholder row

        wb.save(self.output_file)

    def get_rfm_date(self):
        now = datetime.now()
        date = now.date()
        formatted_date = f"cold_{date.day:02d}_{date.month:02d}"
        return formatted_date
    
    def get_date(self):
        now = datetime.now()
        date = now.date()
        formatted_date = f"{date.day:02d}.{date.month:02d}"
        return formatted_date

    def get_total_rows(self):
        """
        Returns the total number of rows across all sheets in the Excel file.
        Used for the final file title renaming.
        """
        wb = load_workbook(self.output_file, read_only=True)  # Use read-only mode for efficiency
        total_rows = sum(wb[sheet].max_row for sheet in wb.sheetnames)  # Sum max_row for all sheets
        wb.close()  # Explicitly close the workbook to release the file handle
        return total_rows - 1