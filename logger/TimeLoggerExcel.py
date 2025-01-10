from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
import os
import time

class TimeLoggerExcel:
    def __init__(self, logs_folder: str = "logs", log_file_name: str = "time_logs.xlsx"):
        """Initializes the logger and ensures the log file exists."""
        self.logs_folder = logs_folder
        self.log_file_path = os.path.join(logs_folder, log_file_name)
        self.start_times = {}
        self.durations = {}

        # Ensures logs folder exists
        os.makedirs(self.logs_folder, exist_ok=True)

        # Creates the Excel file with headers if it doesn't exist
        if not os.path.exists(self.log_file_path):
            self.__create_excel_file()

    def __create_excel_file(self):
        """Creates the initial Excel file with headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Logs"

        # Add headers
        headers = [
            "Date", "Time", "CSV File Rows", "File Name", "Corrupted Unicode",
            "FTD", "Total Parsed Rows", "Parsing Time (Text)", "Parsing Time (Excel)",
            "Processing Time (Text)", "Processing Time (Excel)",
            "Writing Time (Text)", "Writing Time (Excel)"
        ]
        ws.append(headers)

        # Center align the headers
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center")

        wb.save(self.log_file_path)

    def start_timer(self, phase: str):
        """Starts the timer for a given phase (parsing, processing, writing)."""
        self.start_times[phase] = time.perf_counter()

    def stop_timer(self, phase: str):
        """Stops the timer for a given phase and calculates the duration."""
        if phase not in self.start_times:
            raise ValueError(f"Timer for phase '{phase}' was not started.")
        self.durations[phase] = time.perf_counter() - self.start_times[phase]

    def log(self, file_rows: int, file_name: str, corrupted: bool, FTD: str, total_rows: int):
        """Logs the time data into the Excel file."""
        now = datetime.now()
        date = now.strftime("%d.%m.%Y")
        time_ = now.strftime("%H:%M:%S")
        corrupted_str = "YES" if corrupted else "NO"
        FTD = "YES" if FTD == "FTD_" else "NO"

        # Formats times for text and Excel
        parsing_text, parsing_excel = self.__format_time(self.durations.get("parsing", 0.0))
        processing_text, processing_excel = self.__format_time(self.durations.get("processing", 0.0))
        writing_text, writing_excel = self.__format_time(self.durations.get("writing", 0.0))

        # Opens the workbook and append the log
        wb = load_workbook(self.log_file_path)
        ws = wb.active
        ws.append([
            date, time_, file_rows, file_name, corrupted_str, FTD, total_rows,
            parsing_text, parsing_excel,
            processing_text, processing_excel,
            writing_text, writing_excel
        ])

        # Applies Excel time number format to duration columns
        duration_cols = ["I", "K", "M"]  # Columns for Excel durations
        for col in duration_cols:
            ws[f"{col}{ws.max_row}"].number_format = 'hh:mm:ss.000'

        wb.save(self.log_file_path)

    def __format_time(self, seconds: float) -> tuple[str, timedelta]:
        """Formats time for both human-readable text and Excel-compatible duration."""
        # Human-readable format
        ms = int((seconds % 1) * 1000)
        sec = int(seconds % 60)
        minutes = int(seconds // 60)
        text = f"{minutes} min {sec} sec {ms} ms" if minutes > 0 else f"{sec} sec {ms} ms"

        # Excel-compatible duration
        excel_time = timedelta(seconds=seconds)

        return text, excel_time
